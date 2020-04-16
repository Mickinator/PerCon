import pygame.mixer
import RPi.GPIO as GPIO
import time
import openpyxl
import mido
pygame.init()
pygame.mixer.init(44100, 16, 2, 4096)
pygame.mixer.music.set_volume(0.6)
sound = pygame.mixer.Sound('/home/pi/PerCon/audio/mitest.wav')
wb = openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
sheet = wb.get_sheet_by_name("Songs 1")
sheet1 = wb.get_sheet_by_name("Jingles")
sheet2 = wb.get_sheet_by_name("Gimmiks")
sheet3 = wb.get_sheet_by_name("Songs 1")
sheet4 = wb.get_sheet_by_name("Songs 2")
sheet5 = wb.get_sheet_by_name("Songs 3")
sheet6 = wb.get_sheet_by_name("Songs 4")
sheet7 = wb.get_sheet_by_name("Songs 5")
sheet8 = wb.get_sheet_by_name("Songs 6")
sheet9 = wb.get_sheet_by_name("Songs 7")
sheet10 = wb.get_sheet_by_name("Songs 8")
sheet11 = wb.get_sheet_by_name("Knob_Func_1") 
sheet12 = wb.get_sheet_by_name("Knob_Func_2")
sheet13 = wb.get_sheet_by_name("Knob_Func_3")
sheet14 = wb.get_sheet_by_name("Knob_Func_4")
sheet15 = wb.get_sheet_by_name("Knob_Func_5")
sheet16 = wb.get_sheet_by_name("Knob_Func_6")
sheet17 = wb.get_sheet_by_name("Knob_Func_7")
sheet18 = wb.get_sheet_by_name("Knob_Func_8")
sheeta =()
sheetb =()
GPIO.setmode(GPIO.BCM)
GPIO.setup(5, GPIO.IN)  # knob 1
GPIO.setup(22, GPIO.IN) # knob 2
GPIO.setup(27, GPIO.IN) # knob 3
GPIO.setup(17, GPIO.IN) # knob 4
GPIO.setup(4, GPIO.IN)  # knob 5
GPIO.setup(12, GPIO.IN) # knob 6
GPIO.setup(25, GPIO.IN) # knob 7
GPIO.setup(24, GPIO.IN) # knob 8
GPIO.setup(23, GPIO.IN) # knob 9
GPIO.setup(18, GPIO.IN) # knob 10
tog = True # toggle
tom = True # momentary
tof = True # flip
ton = True # single note
toc = True # chords
set = 1    # setzaehler
typ = ()   # Function
knz = ()    # knopfzaehler
z = int()     # zeilenzaehler
knnr = int()
chz = 0
chz1 = True
chz2 = False
chz3 = False
chz4 = False
# ------------------- Tabellen ---------------
def tab(): # liest aus Songs, Spalte M mit der Knopfnummer 
    global typ                     # knnr den Function-Typ.     
    global knz
    global z
    global knnr
    global sheets
    global sheetb
    openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
    sheeta = wb.get_sheet_by_name("Songs " + str(set))
    knnr = sheeta["M" + str(z+1)].value # Knob_Func in Abhängigkeit vom Song
    sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr)) # dito
    typ = str(sheetb["B" + str(knz)].value) # Bestimmung des typs
# ----------------kf_Wahl ----------------
def kf_Wahl(): # waehlt den Function Typ
    global typ
    if typ == 'SI':
        SI()
    if typ == 'TO':
        TO()
    if typ == 'MO':
        MO()
    if typ == 'FL':
        FL()
    if typ == 'NS':
        NS()
    if typ == 'NC':
        NC()
    if typ == 'PC':
        PC()
    if typ == 'SY':
        SY()
    if typ == 'LS':
        LS()
    if typ == 'TR':
        TR()
    if typ == 'CP':
        CP()
# -----------------------------------------
def SI():     # Single CC
    global kz
    global knnr
    global set
    openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
    sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
    CH1 = sheetb["C" + str(knz)].value
    CC1 = sheetb["D" + str(knz)].value
    Val1 = sheetb["F" + str(knz)].value
    with mido.open_output("USB MIDI Interface MIDI 1") as outport:
        outport.send(mido.Message('control_change',        
            channel=CH1, control=CC1, value=Val1)) 
    print('Knob', knz, '=', CH1, CC1, Val1, '    SI')
# -----------------------------------------
def TO():     # CC Toggle
    global knnr
    openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
    sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
    CH1 = sheetb["C" + str(knz)].value
    CC1 = sheetb["D" + str(knz)].value
    Val1 = sheetb["F" + str(knz)].value
    Val2 = sheetb["G" + str(knz)].value
    global tog
    if tog == True:
        with mido.open_output("USB MIDI Interface MIDI 1") as outport:
            outport.send(mido.Message('control_change',        
                channel=CH1, control=CC1, value=Val1))   
            print('Knob 4 =', CH1, CC1, Val1, '    TO 1')
            tog = not tog
    else:
        with mido.open_output("USB MIDI Interface MIDI 1") as outport:
            outport.send(mido.Message('control_change',        
                channel=CH1, control=CC1, value=Val2))
        print('Knob 4 =', CH1, CC1, Val2, '      TO 2')
        tog = not tog
# -----------------------------------------        
def MO():     # CC Momentary
    global knnr
    global knz
    openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
    sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
    CH1 = sheetb["C" + str(knz)].value
    CC1 = sheetb["D" + str(knz)].value
    Val1 = sheetb["F" + str(knz)].value
    Val2 = sheetb["G" + str(knz)].value
    global tom   
    if GPIO.input(4) == 1:        
        if tom == True:
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('control_change',        
                    channel=CH1, control=CC1, value=Val1))   
            print('Knob '+ str(knz), CH1, CC1, Val1, '      MO 1')
            time.sleep(0.3)
            tom = False
    else:        
        if tom == False:
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('control_change',        
                    channel=CH1, control=CC1, value=Val1))   
            print('Knob '+ str(knz), CH1, CC1, Val2, '        MO 2')
            time.sleep(0.3)
            tom = True
# -----------------------------------------    
def FL():     # Channel Flip
    global knnr
    global knz
    openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
    sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
    CH1 = sheetb["C" + str(knz)].value
    CC1 = sheetb["D" + str(knz)].value
    Val1 = sheetb["F" + str(knz)].value
    Val2 = sheetb["G" + str(knz)].value
    CH2 = sheetb["I" + str(knz)].value
    global tof
    if tof == True:
        with mido.open_output("USB MIDI Interface MIDI 1") as outport:
            outport.send(mido.Message('control_change',        
                channel=CH1, control=CC1, value=Val1))
            outport.send(mido.Message('control_change',        
                channel=CH2, control=CC1, value=0))
        print('Knob '+ str(knz), 'CH 1 = ', CH1, CC1, Val1, ' Ch 2 = ', CH2, CC1, Val2, '    FL 1')
        time.sleep(0.5)
        tof = False               
    else:
        with mido.open_output("USB MIDI Interface MIDI 1") as outport:
            outport.send(mido.Message('control_change',        
                channel=CH1, control=CC1, value=0))
            outport.send(mido.Message('control_change',        
                channel=CH2, control=CC1, value=Val2))
        print('Knob '+ str(knz),'Ch 1 = ', CH1, CC1, Val2, ' Ch 2 = ', CH2, CC1, Val1, '    FL 2')
        time.sleep(0.5)
        tof = True        
# -----------------------------------------         
def NS():
    openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
    sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
    CH1 = sheetb["C" + str(knz)].value
    NO = sheetb["K" + str(knz)].value
    Vel = sheetb["L" + str(knz)].value   
    global ton
    if ton == True:
        with mido.open_output("USB MIDI Interface MIDI 1") as outport:
            outport.send(mido.Message('note_on',        
                channel=CH1, note=NO, velocity=Vel))   
        print('Knob '+ str(knz),'=', CH1, NO, Vel, '    NS ON')
        ton = not ton
    else:
        with mido.open_output("USB MIDI Interface MIDI 1") as outport:
            outport.send(mido.Message('note_off',        
                channel=CH1, note=NO, velocity=0))   
        print('Knob '+ str(knz),'=', CH1, NO, '0', '     NS OFF')
        ton = not ton
# ----------------------------------------- 
def NC():
    global knnr
    global knz
    openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
    sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
    CH1 = sheetb["C" + str(knz)].value
    NO1 = sheetb["K" + str(knz)].value
    Vel1 = sheetb["L" + str(knz)].value
    NO2 = sheetb["M" + str(knz)].value
    Vel2 = sheetb["N" + str(knz)].value
    NO3 = sheetb["O" + str(knz)].value
    Vel3 = sheetb["P" + str(knz)].value
    NO4 = sheetb["Q" + str(knz)].value
    Vel4 = sheetb["R" + str(knz)].value
    global toc
    if toc == True:  
        with mido.open_output("USB MIDI Interface MIDI 1") as outport:
            outport.send(mido.Message('note_on',        
                channel=CH1, note=NO1, velocity=Vel1))
            outport.send(mido.Message('note_on',        
                channel=CH1, note=NO2, velocity=Vel2))
            outport.send(mido.Message('note_on',        
                channel=CH1, note=NO3, velocity=Vel3))
            outport.send(mido.Message('note_on',        
                channel=CH1, note=NO4, velocity=Vel4))
        print('Knob '+ str(knz), '=', CH1, NO1, Vel1, '     ', CH1, NO2, Vel2,
               '       ', CH1, NO3, Vel3, '     ', CH1, NO4, Vel4, '    NC ON')
        toc = not toc
    else:
        with mido.open_output("USB MIDI Interface MIDI 1") as outport:
            outport.send(mido.Message('note_off',        
                channel=CH1, note=NO1, velocity=0))
            outport.send(mido.Message('note_off',        
                channel=CH1, note=NO2, velocity=0))
            outport.send(mido.Message('note_off',        
                channel=CH1, note=NO3, velocity=0))
            outport.send(mido.Message('note_off',        
                channel=CH1, note=NO4, velocity=0))
        print('Knob '+ str(knz), '=', CH1, NO1, '0', '      ', CH1, NO2, '0',
               '        ', CH1, NO3, '0', '      ', CH1, NO4, '0', '     NC OFF')
        toc = not toc
# ----------------------------------------- 
def PC():        # Prog Change senden 
    global knnr
    global knz
    global z
    global set
    sheeta = wb.get_sheet_by_name("Songs " + str(set))
    openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
    sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
    with mido.open_output("USB MIDI Interface MIDI 1") as outport:
        outport.send(mido.Message('control_change',
            channel=sheetb["C" + str(knz)].value,
            control=sheetb["D" + str(knz)].value,
            value=sheetb["F" + str(knz)].value))    #   MSB
        time.sleep(0.5)
        outport.send(mido.Message('control_change',
            channel=sheetb["C" + str(knz)].value,
            control=sheetb["E" + str(knz)].value,
            value=sheetb["G" + str(knz)].value))  #    LSB
        time.sleep(0.5)
        outport.send(mido.Message('program_change',
            channel=sheetb["C" + str(knz)].value,
            program=sheetb["H" + str(knz)].value)) #    PC
        time.sleep(0.2)
        outport.send(mido.Message('control_change',
            channel=int(sheeta["G" + str(z+1)].value),
            control=int(sheeta["H" + str(z+1)].value),
            value=int(sheeta["I" + str(z+1)].value)))
        time.sleep(0.5)
      
        print(sheetb["F" + str(knz)].value, sheetb["G" + str(knz)].value,
              sheetb["H" + str(knz)].value)
# ----------------------------------------- 
def SY():
    global knnr
    global knz
    openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
    sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
    syz = 12
    wert = sheetb.cell(row=knz, column=syz).value
    msg = []
    while wert != 247:       
        msg.append(sheetb.cell(row=knz, column=syz).value) # einen dranhaengen
        syz +=1
        wert = (sheetb.cell(row=knz, column=syz).value)
           
    else:
        with mido.open_output("USB MIDI Interface MIDI 1") as outport:
            outport.send(mido.Message('sysex', data=msg))
            print('Knob '+ str(knz), '=', wert, msg, '    SysEx')
# -----------------------------------------             
def CP():
    global knnr
    global knz
    global chz
    global chz1
    global chz2
    global chz3
    global chz4  
    if chz == 0:
        if chz1 == True:
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
            CH1 = sheetb["C" + str(knz)].value
            NO1 = sheetb["W" + str(knz)].value
            NO2 = sheetb["X" + str(knz)].value
            NO3 = sheetb["Y" + str(knz)].value
            NO4 = sheetb["Z" + str(knz)].value
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO1, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO2, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO3, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO4, velocity=0))
            print('Knob '+ str(knz), '=', CH1, NO1, '0', ' ', CH1, NO2, '0',
                   '       ', CH1, NO3, '0', ' ', CH1, NO4, '0')
            CH1 = sheetb["C" + str(knz)].value
            NO1 = sheetb["K" + str(knz)].value
            NO2 = sheetb["L" + str(knz)].value
            NO3 = sheetb["M" + str(knz)].value
            NO4 = sheetb["N" + str(knz)].value
            Vel = sheetb["F" + str(knz)].value
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO1, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO2, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO3, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO4, velocity=Vel))
            print('Knob '+ str(knz), '=', CH1, NO1, Vel, ' ', CH1, NO2, Vel,
                   '       ', CH1, NO3, Vel, ' ', CH1, NO4, Vel, '    CP 1')
            chz1 = False
            chz2 = True
            chz3 = False
            chz4 = False
            time.sleep(0.5)
    if chz == 1:
        if chz2 == True:
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
                            
            CH1 = sheetb["C" + str(knz)].value
            NO1 = sheetb["K" + str(knz)].value
            NO2 = sheetb["L" + str(knz)].value
            NO3 = sheetb["M" + str(knz)].value
            NO4 = sheetb["N" + str(knz)].value
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO1, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO2, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO3, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO4, velocity=0))
            print('Knob '+ str(knz), '=', CH1, NO1, '0', ' ', CH1, NO2, '0',
                   '       ', CH1, NO3, '0', ' ', CH1, NO4, '0')
            CH1 = sheetb["C" + str(knz)].value
            NO1 = sheetb["O" + str(knz)].value
            NO2 = sheetb["P" + str(knz)].value
            NO3 = sheetb["Q" + str(knz)].value
            NO4 = sheetb["R" + str(knz)].value
            Vel = sheetb["F" + str(knz)].value
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO1, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO2, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO3, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO4, velocity=Vel))
            print('Knob '+ str(knz), '=', CH1, NO1, Vel, ' ', CH1, NO2, Vel,
                   '       ', CH1, NO3, Vel, ' ', CH1, NO4, Vel, '    CP 2')
            chz1 = False
            chz2 = False
            chz3 = True
            chz4 = False
            time.sleep(0.5)
    if chz == 2:
        if chz3 == True:
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
            CH1 = sheetb["C" + str(knz)].value
            NO1 = sheetb["O" + str(knz)].value
            NO2 = sheetb["P" + str(knz)].value
            NO3 = sheetb["Q" + str(knz)].value
            NO4 = sheetb["R" + str(knz)].value
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO1, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO2, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO3, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO4, velocity=0))
            print('Knob '+ str(knz), '=', CH1, NO1, '0', ' ', CH1, NO2, '0',
                   '       ', CH1, NO3, '0', ' ', CH1, NO4, '0,')
            CH1 = sheetb["C" + str(knz)].value
            NO1 = sheetb["S" + str(knz)].value
            NO2 = sheetb["T" + str(knz)].value
            NO3 = sheetb["U" + str(knz)].value
            NO4 = sheetb["V" + str(knz)].value
            Vel = sheetb["F" + str(knz)].value
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO1, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO2, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO3, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO4, velocity=Vel))
            print('Knob '+ str(knz), '=', CH1, NO1, Vel, ' ', CH1, NO2, Vel,
                   '       ', CH1, NO3, Vel, ' ', CH1, NO4, Vel, '    CP 3')
            chz1 = False
            chz2 = False
            chz3 = False
            chz4 = True
            time.sleep(0.5) 
    if chz == 3:
        if chz4 == True:
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheetb = wb.get_sheet_by_name("Knob_Func_" + str(knnr))
            CH1 = sheetb["C" + str(knz)].value
            NO1 = sheetb["S" + str(knz)].value
            NO2 = sheetb["T" + str(knz)].value
            NO3 = sheetb["U" + str(knz)].value
            NO4 = sheetb["V" + str(knz)].value
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO1, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO2, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO3, velocity=0))
                outport.send(mido.Message('note_off',        
                    channel=CH1, note=NO4, velocity=0))
            print('Knob '+ str(knz), '=', CH1, NO1, '0', ' ', CH1, NO2, '0',
                   '       ', CH1, NO3, '0', ' ', CH1, NO4, '0')
            CH1 = sheetb["C" + str(knz)].value
            NO1 = sheetb["W" + str(knz)].value
            NO2 = sheetb["X" + str(knz)].value
            NO3 = sheetb["Y" + str(knz)].value
            NO4 = sheetb["Z" + str(knz)].value
            Vel = sheetb["F" + str(knz)].value
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO1, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO2, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO3, velocity=Vel))
                outport.send(mido.Message('note_on',        
                    channel=CH1, note=NO4, velocity=Vel))
            print('Knob '+ str(knz), '=', CH1, NO1, Vel, ' ', CH1, NO2, Vel,
                   '       ', CH1, NO3, Vel, ' ', CH1, NO4, Vel, '    CP 4') 
            chz1 = True
            chz2 = False
            chz3 = False
            chz4 = False
            time.sleep(0.5)
    chz +=1
    if chz == 4:
        chz = 0
# ----------------------------------------- 
def LS():    # Layer Step
    pass
# ----------------------------------------- 
def TR():  # Transpose
    pass
# ----------------------------------------- 
def spielen():
    pygame.mixer.music.play(0)
# ----------------------------------------- 
def spiel2():
    if GPIO.input(22) == 0:
        print()
        print('ja, jetzt ja ! ')
        time.sleep(0.2)
        pygame.mixer.music.play(0)
    else:
        print('noch nicht ! ', end='')
        time.sleep(0.2)
        spiel2()
# -----------------------------------------        
def PerCon():
    global set
    global z
    global knz
    z = 1                                # Zeilenzaehler = SongNr aus Sheet
    gz = 1
    gj = 1
    vpb = 0.6                            # Vol Playback
    shn = 1                              # Sheetnumber
    vzy = sheet3["I" + str(z+1)].value   # Vol Zynthian aus Sheet1
    sz = (sheet3["C" + str(z+1)].value)  # Name des Songs
    sn = (sheet3["A" + str(z+1)].value)  # Songnr aus Sheet
    print()
    print(' --- Mickis Performance Controller ---')
    print()
    print(' --- Setwahl-Modus ---')
    print()
    print(' --- Bitte Set waehlen ---')
    print(' --- Knopf 1 - 8 fuer Sets 1 - 8 ---')
    print(' --- Knopf 9 fuer ENDE ---')
    print(' --- Knopf 10 fuer START ---')
    pygame.mixer.music.load('/home/pi/PerCon/sets/choose_set.mp3')
    spielen()
    while GPIO.input(18) != 0:
# -------------------- Setwahl ----------------------
# ---------------------Set 1 ---------------------------
        if GPIO.input(5) == 0:         
            pygame.mixer.music.load('/home/pi/PerCon/sets/set1.mp3')
            set = 1
            print('Set 1')
            time.sleep(0.5)
            spielen()
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheet = wb.get_sheet_by_name("Songs 1")
            shn = sheet3
            sz = sheet3["C" + str(z+1)].value
            sn = sheet3["A" + str(z+1)].value
# ---------------------Set 2 ---------------------------            
        if GPIO.input(22) == 0:                
            pygame.mixer.music.load('/home/pi/PerCon/sets/set2.mp3')
            set = 2
            print('Set 2')
            time.sleep(0.5)
            spielen()
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheet = wb.get_sheet_by_name("Songs 2")  
            shn = sheet4
            sz = sheet4["C" + str(z+1)].value
            sn = sheet4["A" + str(z+1)].value
# ---------------------Set 3 ---------------------------
        if GPIO.input(27) == 0:                       
            pygame.mixer.music.load('/home/pi/PerCon/sets/set3.mp3')
            set = 3
            print('Set 3')
            time.sleep(0.5)
            spielen()
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheet = wb.get_sheet_by_name("Songs 3")  
            shn = sheet5
            sz = sheet5["C" + str(z+1)].value
            sn = sheet5["A" + str(z+1)].value
# ---------------------Set 4 ---------------------------
        if GPIO.input(17) == 0:                  
            pygame.mixer.music.load('/home/pi/PerCon/sets/set4.mp3')
            set = 4
            print('Set 4')
            time.sleep(0.5)
            spielen()
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheet = wb.get_sheet_by_name("Songs 4")
            shn = sheet6
            sz = sheet6["C" + str(z+1)].value
            sn = sheet6["A" + str(z+1)].value
# ---------------------Set 5 ---------------------------
        if GPIO.input(4) == 0:              
            pygame.mixer.music.load('/home/pi/PerCon/sets/set5.mp3')
            set = 5
            print('Set 5')
            time.sleep(0.5)
            spielen()
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheet = wb.get_sheet_by_name("Songs 5")          
            shn = sheet7
            sz = sheet7["C" + str(z+1)].value
            sn = sheet7["A" + str(z+1)].value
# ---------------------Set 6 ---------------------------
        if GPIO.input(12) == 0:      
            pygame.mixer.music.load('/home/pi/PerCon/sets/set6.mp3')
            set = 6
            print('Set 6')
            time.sleep(0.5)
            spielen()
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheet = wb.get_sheet_by_name("Songs 6")        
            shn = sheet8
            sz = sheet8["C" + str(z+1)].value
            sn = sheet8["A" + str(z+1)].value
# ---------------------Set 7 ---------------------------
        if GPIO.input(25) == 0:                   
            pygame.mixer.music.load('/home/pi/PerCon/sets/set7.mp3')
            set = 7
            print('Set 7')
            time.sleep(0.5)
            spielen()
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheet = wb.get_sheet_by_name("Songs 7")     
            shn = sheet9
            sz = sheet9["C" + str(z+1)].value
            sn = sheet9["A" + str(z+1)].value
# ---------------------Set 8 ---------------------------
        if GPIO.input(24) == 0:                 
            pygame.mixer.music.load('/home/pi/PerCon/sets/set8.mp3')
            set = 8
            print('Set 8')
            time.sleep(0.5)
            spielen()
            openpyxl.load_workbook("/home/pi/PerCon/data/songs.xlsx")
            sheet = wb.get_sheet_by_name("Songs 8")   
            shn = sheet10
            sz = sheet10["C" + str(z+1)].value
            sn = sheet10["A" + str(z+1)].value
# ---------------------raus---------------------------
        if GPIO.input(23) == 0:
            print()
            print(' --- Mickis Performance Controller ---')
            print()
            print(' --- Ende der Sendung ---')
            time.sleep(0.5)
            exit()
    # ------------- Setlist -----------------------------
    pygame.mixer.music.load('/home/pi/PerCon/sets/jetzt.mp3')
    print(' ---- Jetzt gehts los ! -----')
    print()
    print(' --- Songwahl-Modus ---')
    print(' --- Bitte Knopf 1 für vorwärts  ---')
    print(' --- Bitte Knopf 2 für rückwärts ---')
    print(' --- Knopf 9 fuer EXIT zur Setwahl ---')
    print()
    pygame.mixer.music.load('/home/pi/PerCon/sets/start_song.mp3')
    spielen()                         # --Jetzt gehts los ---
    time.sleep(0.5)
# -----------------------Songwahl -------------------        
    while True:  
# ----------------------- Knob 1 -vorwaerts----------
      
        if GPIO.input(5) == 0:    # Song +1       verwendet D E F
            z += 1                     # Zeile 3
            sz = shn["C" + str(z+1)].value
            if sz == str('ENDE'):
                print(' ---  Ende der Setlist ---')
                print(' --- Bitte Set waehlen ---')
                pygame.mixer.music.load('/home/pi/PerCon/sets/wrong_song.mp3')
                spielen()
                time.sleep(0.5)
                PerCon()
            sn = shn["A" + str(z+1)].value
            vzy = sheet["I" + str(z+1)].value   # Vol Zynthian aus Sheet1
            print()
            print('SongNr=', sn, sz)            # Songnummer, Titel
            time.sleep(0.5)
            pygame.mixer.music.load('/home/pi/PerCon/audio/' + sz)  # akt Songladen        
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:

# Prog Change senden                
                outport.send(mido.Message('control_change',
                    channel=15,
                    control=0,
                    value=shn["E" + str(z+1)].value))    #   MSB         
                outport.send(mido.Message('control_change',
                    channel=15,
                    control=32,
                    value=shn["F" + str(z+1)].value))  #    LSB                 
                outport.send(mido.Message('program_change',
                    channel=15,
                    program=shn["D" + str(z+1)].value)) #    PC                         
                print('PC = ', shn["E" + str(z+1)].value,
                      shn["F" + str(z+1)].value,
                      shn["D" + str(z+1)].value)
# Vol Zynthian senden                
                outport.send(mido.Message('control_change',        
                    channel=0, control=7, value=shn["I" + str(z+1)].value))   # Start Volume Zynthian         
                print('Vol Zyn = ', shn["I" + str(z+1)].value)         
# Vol Playback senden                
                pygame.mixer.music.set_volume(shn["L" + str(z+1)].value)      # Start Volume Playback                
                print('Vol Pb  = ', shn["L" + str(z+1)].value)          
# ---------akt Song spielen -----------------------           
            spiel2()
            time.sleep(0.5)
# ----------------------- Knob 2 --rueckwaerts-----------------                   
        if GPIO.input(22) == 0:   # Song -1      verwendet verwendet D E F
            z -= 1
            sz = shn["C" + str(z+1)].value
            if sz == str('Name'):
                print(' ---  Anfang der Setlist ---')
                print(' --- Bitte Set waehlen ---')
                pygame.mixer.music.load('/home/pi/PerCon/sets/wrong_song.mp3')
                spielen()
                PerCon()
            sz = shn["C" + str(z+1)].value
            if sz == str('Name'):
                print(' --- Bitte Set waehlen ---')
                pygame.mixer.music.load('/home/pi/PerCon/sets/wrong_song.mp3')
                spielen()
                PerCon()
            sz = shn["C" + str(z+1)].value
            sn = shn["A" + str(z+1)].value
            vzy = shn["I" + str(z+1)].value          # Vol Zynthian aus Sheet1
            print()
            print('SongNr=', sn, sz)
            pygame.mixer.music.load('/home/pi/PerCon/audio/' + sz)
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('control_change',
                    channel=15,
                    control=0,
                    value=shn["E" + str(z+2)].value))   # MSB        
                outport.send(mido.Message('control_change', channel=15, control=32, value=sheet["F" + str(z+1)].value))  # LSB  
                outport.send(mido.Message('program_change',
                    channel=15, program=sheet["D" + str(z+1)].value))            # PC  
                print(shn["E" + str(z+1)].value,
                      shn["F" + str(z+1)].value,
                      shn["D" + str(z+1)].value)
                outport.send(mido.Message('control_change',        
                    channel=0, control=7, value=shn["I" + str(z+1)].value))   # Start Volume Zynthian   
                print('Vol Zyn = ', shn["I" + str(z+1)].value)
                pygame.mixer.music.set_volume(shn["L" + str(z+1)].value)      # Start Volume Playback   
                print('Vol Pb  = ', shn["L" + str(z+1)].value)       
            spiel2() #akt Song spielen
            time.sleep(0.5)
            print()
#-------------------Knob 3 --------------------------------------------        
        if GPIO.input(27) == 0:
            knz = 3
            tab()
            kf_Wahl()        
# ------- Knob 4 -------------------------------------------------          
        if GPIO.input(17) == 0:
            knz = 4
            tab()
            kf_Wahl()
# ------- Knob 5 -------------------------------------------------
        if GPIO.input(4) == 0:
            knz = 5
            tab()
            kf_Wahl()
# --------Knob 6 -------------------------------------------------
        if GPIO.input(12) == 0:
            knz = 6
            tab()
            kf_Wahl()
# --------Knob 7 -------------------------------------------------
        if GPIO.input(25) == 0:
            knz = 7
            time.sleep(0.5)
            tab()
            kf_Wahl()
# --------Knob 8 -------------------------------------------------                
        if GPIO.input(24) == 0:
            knz = 8
            tab()
            kf_Wahl()
# --------Knob 9 -------------------------------------------------
        if GPIO.input(23) == 0:             
            print(' --- zurueck zur Setwahl ---')
            print()
            print(' --- Setwahl-Modus ---')
            print(' --- Knopf 1 - 8 fuer Sets 1 - 8 ---')
            print(' --- Knopf 9 fuer ENDE ---')
            pygame.mixer.music.load('/home/pi/PerCon/sets/choose_set.mp3')
            spielen()
            time.sleep(0.5)
            pygame.mixer.pause()
            PerCon()       
    #------------------ Knob 10 ---------------------------
        if GPIO.input(18) == 0:  
            with mido.open_output("USB MIDI Interface MIDI 1") as outport:
                outport.send(mido.Message('control_change',
                    channel=0, control=123, value=0))  
                time.sleep(0.5)
                print('Knopf 10: ANO')
# ---------------------------------------------------------
PerCon()